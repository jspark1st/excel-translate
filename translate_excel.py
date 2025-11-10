"""
엑셀 파일 번역 프로그램
구글 번역 API를 사용하여 엑셀 파일의 내용을 한국어로 번역합니다.
"""

import pandas as pd
import os
import sys
from googletrans import Translator
from typing import Union
import time
import traceback
import logging
from datetime import datetime
from logger_config import setup_logger
from httpcore._exceptions import ReadTimeout
from httpx import Timeout
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from copy import copy

# 로거 설정 (기본값: INFO 레벨로 설정하여 성능 향상)
# DEBUG 모드는 필요시 코드에서 직접 활성화
logger = setup_logger("translate_excel", logging.INFO)


class ExcelTranslator:
    """엑셀 파일 번역 클래스"""
    
    def __init__(self, debug_mode: bool = False, progress_callback=None):
        """
        번역기 초기화
        
        Args:
            debug_mode: 디버그 모드 활성화 여부
            progress_callback: 진행률 업데이트 콜백 함수 (current, total, detail) -> None
        """
        logger.debug("ExcelTranslator 초기화 시작")
        self.debug_mode = debug_mode
        self.start_time = datetime.now()
        self.progress_callback = progress_callback  # 진행률 콜백 함수
        
        try:
            # 타임아웃 설정 (프록시 환경 고려하여 증가)
            # connect_timeout: 연결 타임아웃, read_timeout: 읽기 타임아웃, write_timeout: 쓰기 타임아웃, pool_timeout: 풀 타임아웃
            timeout = Timeout(
                connect_timeout=10.0,
                read_timeout=30.0,
                write_timeout=10.0,
                pool_timeout=5.0
            )
            self.translator = Translator(timeout=timeout)
            logger.info("구글 번역 API 초기화 완료 (타임아웃: 연결 10초, 읽기 30초)")
        except Exception as e:
            logger.error(f"번역기 초기화 실패: {str(e)}", exc_info=True)
            raise
        
        self.translated_count = 0
        self.error_count = 0
        self.skipped_count = 0  # 건너뛴 셀 수 (숫자, 빈 셀 등)
        self.total_cells = 0
        self.total_cells_to_process = 0  # 처리할 전체 셀 수 (진행률 계산용)
        self.last_progress_update = datetime.now()  # 마지막 진행률 업데이트 시간 (성능 최적화용)
        self.should_stop = False  # 번역 중지 플래그
        
        logger.debug("ExcelTranslator 초기화 완료")
    
    def _translate_header_cell(self, text: str) -> str:
        """
        헤더 셀 전용 번역 메서드 (숫자 체크 건너뛰기)
        
        Args:
            text: 번역할 텍스트 (문자열)
            
        Returns:
            번역된 텍스트 (한국어)
        """
        try:
            # 빈 문자열 체크
            if not text or not text.strip():
                return text
            
            text_str = text.strip()
            
            # 이미 한국어인지 체크
            if any('\uAC00' <= char <= '\uD7A3' for char in text_str[:20]):
                return text_str
            
            # 구글 번역 API 호출 (재시도 로직 포함)
            translate_start = datetime.now()
            
            # 재시도 로직 (최대 3회, 지수 백오프)
            max_retries = 3
            retry_delays = [1, 2, 4]  # 1초, 2초, 4초
            result = None
            last_exception = None
            
            for attempt in range(max_retries):
                try:
                    result = self.translator.translate(text_str, dest='ko')
                    break  # 성공 시 루프 종료
                except ReadTimeout as e:
                    last_exception = e
                    if attempt < max_retries - 1:
                        wait_time = retry_delays[attempt]
                        logger.warning(
                            f"헤더 번역 타임아웃 (시도 {attempt + 1}/{max_retries}): "
                            f"'{text_str[:50]}...' - {wait_time}초 후 재시도"
                        )
                        time.sleep(wait_time)
                    else:
                        logger.error(
                            f"헤더 번역 타임아웃 (최대 재시도 횟수 초과): '{text_str[:50]}...'"
                        )
                except Exception as e:
                    last_exception = e
                    if isinstance(e, (ConnectionError, TimeoutError)) and attempt < max_retries - 1:
                        wait_time = retry_delays[attempt]
                        logger.warning(
                            f"헤더 네트워크 오류 (시도 {attempt + 1}/{max_retries}): "
                            f"'{text_str[:50]}...' - {wait_time}초 후 재시도"
                        )
                        time.sleep(wait_time)
                    else:
                        raise
            
            if result is None:
                raise last_exception if last_exception else Exception("헤더 번역 실패")
            
            translate_duration = (datetime.now() - translate_start).total_seconds()
            self.translated_count += 1
            
            # googletrans는 무료 API이지만 과도한 딜레이는 불필요
            time.sleep(0.01)
            
            return result.text
            
        except ReadTimeout:
            logger.error(f"헤더 번역 타임아웃: {text_str[:50] if 'text_str' in locals() else str(text)[:50]}...")
            return text_str if 'text_str' in locals() else str(text)
        except Exception as e:
            logger.error(f"헤더 번역 오류: {text_str[:50] if 'text_str' in locals() else str(text)[:50]}... - {str(e)}")
            return text_str if 'text_str' in locals() else str(text)
    
    def translate_text(self, text: Union[str, float, int]) -> str:
        """
        텍스트를 한국어로 번역
        
        Args:
            text: 번역할 텍스트
            
        Returns:
            번역된 텍스트 (한국어)
        """
        self.total_cells += 1
        
        try:
            # 숫자나 None 값은 그대로 반환
            if pd.isna(text) or text == '':
                self.skipped_count += 1
                if self.debug_mode:
                    logger.debug(f"빈 셀 건너뜀: {text}")
                # 진행률 업데이트 (성능 최적화: 10개마다 또는 0.5초마다)
                if self.progress_callback and self.total_cells_to_process > 0:
                    processed = self.translated_count + self.skipped_count
                    if processed % 10 == 0 or (datetime.now() - self.last_progress_update).total_seconds() > 0.5:
                        self.progress_callback(processed, self.total_cells_to_process)
                        self.last_progress_update = datetime.now()
                return text
            
            # 숫자 타입은 문자열로 변환하지 않고 그대로 반환
            if isinstance(text, (int, float)):
                self.skipped_count += 1
                if self.debug_mode:
                    logger.debug(f"숫자 셀 건너뜀: {text}")
                # 진행률 업데이트 (성능 최적화: 10개마다 또는 0.5초마다)
                if self.progress_callback and self.total_cells_to_process > 0:
                    processed = self.translated_count + self.skipped_count
                    if processed % 10 == 0 or (datetime.now() - self.last_progress_update).total_seconds() > 0.5:
                        self.progress_callback(processed, self.total_cells_to_process)
                        self.last_progress_update = datetime.now()
                return text
            
            text_str = str(text).strip()
            
            # 빈 문자열이면 그대로 반환
            if not text_str:
                self.skipped_count += 1
                if self.debug_mode:
                    logger.debug("빈 문자열 건너뜀")
                # 진행률 업데이트 (성능 최적화: 10개마다 또는 0.5초마다)
                if self.progress_callback and self.total_cells_to_process > 0:
                    processed = self.translated_count + self.skipped_count
                    if processed % 10 == 0 or (datetime.now() - self.last_progress_update).total_seconds() > 0.5:
                        self.progress_callback(processed, self.total_cells_to_process)
                        self.last_progress_update = datetime.now()
                return text
            
            # 이미 한국어인지 간단 체크 (한글 포함 여부) - 최적화된 버전
            # 처음 20자만 체크하여 성능 향상 (대부분의 경우 충분)
            if any('\uAC00' <= char <= '\uD7A3' for char in text_str[:20]):
                self.skipped_count += 1
                if self.debug_mode:
                    logger.debug(f"이미 한글 포함 텍스트 건너뜀: {text_str[:50]}")
                # 진행률 업데이트 (성능 최적화: 10개마다 또는 0.5초마다)
                if self.progress_callback and self.total_cells_to_process > 0:
                    processed = self.translated_count + self.skipped_count
                    if processed % 10 == 0 or (datetime.now() - self.last_progress_update).total_seconds() > 0.5:
                        self.progress_callback(processed, self.total_cells_to_process)
                        self.last_progress_update = datetime.now()
                return text_str
            
            # 구글 번역 API 호출 (재시도 로직 포함)
            if self.debug_mode:
                logger.debug(f"번역 시작: {text_str[:100]}")
            translate_start = datetime.now()
            
            # 재시도 로직 (최대 3회, 지수 백오프)
            max_retries = 3
            retry_delays = [1, 2, 4]  # 1초, 2초, 4초
            result = None
            last_exception = None
            
            for attempt in range(max_retries):
                try:
                    result = self.translator.translate(text_str, dest='ko')
                    break  # 성공 시 루프 종료
                except ReadTimeout as e:
                    last_exception = e
                    if attempt < max_retries - 1:
                        wait_time = retry_delays[attempt]
                        logger.warning(
                            f"번역 타임아웃 (시도 {attempt + 1}/{max_retries}): "
                            f"'{text_str[:50]}...' - {wait_time}초 후 재시도"
                        )
                        time.sleep(wait_time)
                    else:
                        logger.error(
                            f"번역 타임아웃 (최대 재시도 횟수 초과): '{text_str[:50]}...'"
                        )
                except Exception as e:
                    # 타임아웃이 아닌 다른 예외는 즉시 재시도하지 않음
                    last_exception = e
                    if isinstance(e, (ConnectionError, TimeoutError)) and attempt < max_retries - 1:
                        wait_time = retry_delays[attempt]
                        logger.warning(
                            f"네트워크 오류 (시도 {attempt + 1}/{max_retries}): "
                            f"'{text_str[:50]}...' - {wait_time}초 후 재시도"
                        )
                        time.sleep(wait_time)
                    else:
                        raise  # 다른 예외는 즉시 전파
            
            # 재시도 후에도 실패한 경우 예외 발생
            if result is None:
                raise last_exception if last_exception else Exception("번역 실패")
            
            translate_duration = (datetime.now() - translate_start).total_seconds()
            self.translated_count += 1
            
            # DEBUG 로그는 디버그 모드에서만 기록 (성능 향상)
            if self.debug_mode:
                logger.debug(
                    f"번역 완료 ({translate_duration:.2f}초): "
                    f"'{text_str[:50]}' -> '{result.text[:50]}'"
                )
            elif self.translated_count % 10 == 0:  # 10개마다 INFO 로그
                logger.info(f"번역 진행 중... ({self.translated_count}개 완료)")
            
            # 진행률 콜백 호출 (성능 최적화: 10개마다 또는 0.5초마다)
            if self.progress_callback and self.total_cells_to_process > 0:
                processed = self.translated_count + self.skipped_count
                if processed % 10 == 0 or (datetime.now() - self.last_progress_update).total_seconds() > 0.5:
                    self.progress_callback(processed, self.total_cells_to_process, 
                                         f"번역 중... ({self.translated_count}개 번역 완료)")
                    self.last_progress_update = datetime.now()
            
            # API 제한을 고려한 딜레이 (최소화)
            # googletrans는 무료 API이지만 과도한 딜레이는 불필요
            time.sleep(0.01)  # 0.1초 → 0.01초로 최적화 (10배 향상)
            
            return result.text
            
        except ReadTimeout as e:
            # 타임아웃 예외 특별 처리
            self.error_count += 1
            error_msg = f"번역 타임아웃 (재시도 실패): {text_str[:50] if 'text_str' in locals() else str(text)[:50]}..."
            logger.error(error_msg, exc_info=True)
            
            if self.debug_mode:
                print(error_msg)
            
            # 타임아웃 발생 시 원문 반환 (재시도 후에도 실패한 경우)
            return text_str if 'text_str' in locals() else str(text)
            
        except Exception as e:
            # 기타 예외 처리
            self.error_count += 1
            error_msg = f"번역 오류 (원문: {text_str[:50] if 'text_str' in locals() else str(text)[:50]}...): {str(e)}"
            logger.error(error_msg, exc_info=True)
            logger.debug(f"스택 트레이스:\n{traceback.format_exc()}")
            
            if self.debug_mode:
                print(error_msg)
            
            return text_str if 'text_str' in locals() else str(text)  # 오류 발생 시 원문 반환
    
    def translate_dataframe(self, df: pd.DataFrame, sheet_name: str = "") -> pd.DataFrame:
        """
        DataFrame의 모든 셀을 번역
        
        Args:
            df: 번역할 DataFrame
            sheet_name: 시트 이름 (로깅용)
            
        Returns:
            번역된 DataFrame
        """
        logger.info(f"DataFrame 번역 시작 (시트: {sheet_name}, 행: {len(df)}, 열: {len(df.columns)})")
        translated_df = df.copy()
        
        start_time = datetime.now()
        total_cells_before = self.total_cells
        
        print("번역 중...")
        logger.debug(f"컬럼 목록: {list(df.columns)}")
        
        # 컬럼명(헤더) 번역
        translated_columns = {}
        logger.info("컬럼명(헤더) 번역 시작")
        for col in df.columns:
            if self.should_stop:
                raise InterruptedError("번역이 사용자에 의해 중지되었습니다.")
            translated_col = self.translate_text(col)
            translated_columns[col] = translated_col
            logger.debug(f"컬럼명 번역: '{col}' -> '{translated_col}'")
        
        # 번역된 컬럼명으로 DataFrame 컬럼명 변경
        translated_df.columns = [translated_columns.get(col, col) for col in df.columns]
        logger.info("컬럼명(헤더) 번역 완료")
        
        for col_idx, col in enumerate(df.columns, 1):
            # 중지 플래그 확인
            if self.should_stop:
                logger.warning(f"번역 중지 요청됨 (시트: {sheet_name}, 컬럼: {col})")
                raise InterruptedError("번역이 사용자에 의해 중지되었습니다.")
            
            col_start_time = datetime.now()
            col_translated_before = self.translated_count
            logger.info(f"컬럼 '{col}' 번역 시작 ({col_idx}/{len(df.columns)})")
            print(f"  컬럼 '{col}' 번역 중... ({col_idx}/{len(df.columns)})")
            
            try:
                # 진행 상황 표시를 위한 콜백 함수
                def translate_with_progress(cell_value):
                    # 중지 플래그 확인
                    if self.should_stop:
                        raise InterruptedError("번역이 사용자에 의해 중지되었습니다.")
                    result = self.translate_text(cell_value)
                    # 10개마다 진행 상황 출력
                    if self.translated_count > 0 and self.translated_count % 10 == 0:
                        print(f"    진행 중... ({self.translated_count}개 번역 완료)", end='\r')
                    return result
                
                # 번역된 컬럼명 사용
                translated_col_name = translated_columns[col]
                translated_df[translated_col_name] = df[col].apply(translate_with_progress)
                col_duration = (datetime.now() - col_start_time).total_seconds()
                col_translated = self.translated_count - col_translated_before
                
                logger.info(
                    f"컬럼 '{col}' 번역 완료 ({col_duration:.2f}초, "
                    f"번역된 셀: {col_translated}/{len(df)}, 누적 번역: {self.translated_count})"
                )
                print(f"  ✓ 컬럼 '{col}' 완료 ({col_translated}개 번역, {col_duration:.2f}초)")
            except InterruptedError:
                # 중지 요청은 예외로 전파하여 상위에서 처리
                raise
            except Exception as e:
                logger.error(f"컬럼 '{col}' 번역 중 오류 발생: {str(e)}", exc_info=True)
                logger.debug(f"스택 트레이스:\n{traceback.format_exc()}")
                raise
        
        duration = (datetime.now() - start_time).total_seconds()
        logger.info(f"DataFrame 번역 완료 (소요 시간: {duration:.2f}초)")
        
        return translated_df
    
    def _copy_sheet_with_formatting(self, source_ws, output_ws, translated_df, sheet_name):
        """
        원본 시트의 서식을 복사하면서 번역된 데이터를 저장
        
        Args:
            source_ws: 원본 워크시트 (openpyxl)
            output_ws: 출력 워크시트 (openpyxl)
            translated_df: 번역된 DataFrame
            sheet_name: 시트 이름
        """
        logger.debug(f"서식 복사 시작 (시트: {sheet_name})")
        
        # 원본 파일의 모든 행과 열 복사 (서식 완벽 유지)
        max_row = source_ws.max_row  # 원본 파일의 모든 행
        max_col = source_ws.max_column  # 원본 파일의 모든 열
        
        # 모든 컬럼 너비 복사
        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            if col_letter in source_ws.column_dimensions:
                source_width = source_ws.column_dimensions[col_letter].width
                if source_width:
                    output_ws.column_dimensions[col_letter].width = source_width
        
        # 모든 행과 셀의 서식 복사
        for row_idx in range(1, max_row + 1):
            # 중지 플래그 확인
            if self.should_stop:
                raise InterruptedError("번역이 사용자에 의해 중지되었습니다.")
            
            # 행 높이 복사
            if row_idx in source_ws.row_dimensions:
                source_height = source_ws.row_dimensions[row_idx].height
                if source_height:
                    output_ws.row_dimensions[row_idx].height = source_height
            
            for col_idx in range(1, max_col + 1):
                col_letter = get_column_letter(col_idx)
                source_cell = source_ws[f"{col_letter}{row_idx}"]
                output_cell = output_ws[f"{col_letter}{row_idx}"]
                
                # 셀 서식 복사 (스타일, 폰트, 테두리, 채우기 등) - 모든 셀에 대해
                try:
                    if source_cell.has_style:
                        if source_cell.font:
                            output_cell.font = copy(source_cell.font)
                        if source_cell.border:
                            output_cell.border = copy(source_cell.border)
                        if source_cell.fill:
                            output_cell.fill = copy(source_cell.fill)
                        if source_cell.number_format:
                            output_cell.number_format = source_cell.number_format
                        if source_cell.protection:
                            output_cell.protection = copy(source_cell.protection)
                        if source_cell.alignment:
                            output_cell.alignment = copy(source_cell.alignment)
                except Exception as e:
                    logger.debug(f"셀 서식 복사 중 오류 (무시): {str(e)}")
                
                # 번역된 데이터 쓰기 (데이터가 있는 부분만)
                if row_idx == 1:
                    # 헤더 행 - 원본 파일의 첫 행을 직접 읽어서 번역
                    source_value = source_cell.value
                    if source_value is not None:
                        # 헤더 행의 경우 숫자도 문자열로 변환하여 번역 시도
                        # 숫자로 시작하는 헤더(예: "1月份")도 번역되도록 처리
                        if isinstance(source_value, (int, float)):
                            # 순수 숫자인 경우 문자열로 변환
                            source_value_str = str(source_value)
                        else:
                            source_value_str = str(source_value).strip()
                        
                        # 빈 문자열이 아니면 번역 시도
                        if source_value_str:
                            # 헤더는 항상 번역 시도 (숫자 체크 건너뛰기)
                            translated_header_value = self._translate_header_cell(source_value_str)
                            output_cell.value = translated_header_value
                            logger.debug(f"헤더 셀 번역: '{source_value}' -> '{translated_header_value}'")
                        else:
                            output_cell.value = None
                    else:
                        # 빈 셀은 그대로 유지
                        output_cell.value = None
                else:
                    # 데이터 행
                    df_row_idx = row_idx - 2  # 헤더(1) 제외하고 0부터 시작
                    if 0 <= df_row_idx < len(translated_df) and col_idx <= len(translated_df.columns):
                        # 번역된 데이터가 있는 경우
                        col_name = translated_df.columns[col_idx - 1]
                        value = translated_df.iloc[df_row_idx, col_idx - 1]
                        output_cell.value = value
                    else:
                        # 번역 범위를 벗어난 셀은 원본 값 유지
                        output_cell.value = source_cell.value
        
        # 병합된 셀 복사 (원본 파일의 모든 병합 셀)
        if hasattr(source_ws, 'merged_cells') and source_ws.merged_cells:
            for merged_range in list(source_ws.merged_cells.ranges):
                try:
                    output_ws.merge_cells(str(merged_range))
                except Exception as e:
                    logger.debug(f"병합 셀 복사 실패 (무시): {str(e)}")
        
        logger.debug(f"서식 복사 완료 (시트: {sheet_name})")
    
    def translate_excel(self, input_file: str, output_file: str = None):
        """
        엑셀 파일을 번역하여 새 파일로 저장
        
        Args:
            input_file: 입력 엑셀 파일 경로
            output_file: 출력 엑셀 파일 경로 (None이면 자동 생성)
        """
        logger.info("=" * 80)
        logger.info("엑셀 파일 번역 시작")
        logger.info(f"입력 파일: {input_file}")
        
        try:
            # 파일 존재 확인
            if not os.path.exists(input_file):
                error_msg = f"파일을 찾을 수 없습니다: {input_file}"
                logger.error(error_msg)
                raise FileNotFoundError(error_msg)
            
            # 파일 크기 확인
            file_size = os.path.getsize(input_file)
            logger.info(f"파일 크기: {file_size / 1024 / 1024:.2f} MB")
            
            # 출력 파일명 자동 생성
            if output_file is None:
                base_name = os.path.splitext(input_file)[0]
                output_file = f"{base_name}_translated.xlsx"
            
            logger.info(f"출력 파일: {output_file}")
            print(f"입력 파일: {input_file}")
            print(f"출력 파일: {output_file}")
            print("-" * 50)
            
            # 엑셀 파일 읽기 (모든 시트 포함)
            logger.debug("엑셀 파일 열기 시작")
            try:
                excel_file = pd.ExcelFile(input_file)
                sheet_names = excel_file.sheet_names
                logger.info(f"엑셀 파일 열기 성공, 시트 수: {len(sheet_names)}")
                logger.debug(f"시트 목록: {sheet_names}")
            except Exception as e:
                logger.error(f"엑셀 파일 읽기 실패: {str(e)}", exc_info=True)
                raise
            
            print(f"총 {len(sheet_names)}개의 시트를 발견했습니다.")
            
            # 전체 셀 수 계산 (진행률 계산용)
            total_cells_count = 0
            for sheet_name in sheet_names:
                df_temp = pd.read_excel(excel_file, sheet_name=sheet_name)
                total_cells_count += len(df_temp) * len(df_temp.columns)
            self.total_cells_to_process = total_cells_count
            logger.info(f"전체 처리할 셀 수: {total_cells_count}")
            
            # 진행률 초기화
            if self.progress_callback:
                self.progress_callback(0, total_cells_count, "파일 분석 완료, 번역 시작...")
                self.last_progress_update = datetime.now()
            
            # openpyxl로 원본 파일 열기 (서식 유지를 위해)
            logger.debug("원본 파일을 openpyxl로 열기 시작")
            source_wb = load_workbook(input_file)
            logger.info("원본 파일 열기 완료 (서식 유지 모드)")
            
            # 새 워크북 생성 (출력용)
            from openpyxl import Workbook
            output_wb = Workbook()
            output_wb.remove(output_wb.active)  # 기본 시트 제거
            
            # ExcelWriter로 모든 시트를 번역하여 저장
            for sheet_idx, sheet_name in enumerate(sheet_names, 1):
                # 중지 플래그 확인
                if self.should_stop:
                    logger.warning(f"번역 중지 요청됨 (시트: {sheet_name})")
                    raise InterruptedError("번역이 사용자에 의해 중지되었습니다.")
                
                sheet_start_time = datetime.now()
                logger.info(f"[{sheet_idx}/{len(sheet_names)}] 시트 '{sheet_name}' 처리 시작")
                print(f"\n시트 '{sheet_name}' 처리 중...")
                
                try:
                    # 원본 시트와 데이터 읽기
                    logger.debug(f"시트 '{sheet_name}' 읽기 시작")
                    source_ws = source_wb[sheet_name]
                    df = pd.read_excel(excel_file, sheet_name=sheet_name)
                    logger.info(f"시트 '{sheet_name}' 읽기 완료 (행: {len(df)}, 열: {len(df.columns)})")
                    print(f"  행 수: {len(df)}, 열 수: {len(df.columns)}")
                    
                    # 번역
                    translated_df = self.translate_dataframe(df, sheet_name=sheet_name)
                    
                    # 새 시트 생성
                    output_ws = output_wb.create_sheet(title=sheet_name)
                    
                    # 번역된 데이터와 서식 복사
                    logger.debug(f"시트 '{sheet_name}' 저장 시작 (서식 복사 포함)")
                    self._copy_sheet_with_formatting(source_ws, output_ws, translated_df, sheet_name)
                    
                    sheet_duration = (datetime.now() - sheet_start_time).total_seconds()
                    logger.info(f"시트 '{sheet_name}' 처리 완료 (소요 시간: {sheet_duration:.2f}초)")
                    print(f"  시트 '{sheet_name}' 번역 완료")
                    
                except InterruptedError:
                    # 중지 요청은 예외로 전파하여 상위에서 처리
                    raise
                except Exception as e:
                    logger.error(f"시트 '{sheet_name}' 처리 중 오류: {str(e)}", exc_info=True)
                    logger.debug(f"스택 트레이스:\n{traceback.format_exc()}")
                    raise
            
            # 워크북 저장
            logger.debug("번역된 파일 저장 시작")
            output_wb.save(output_file)
            source_wb.close()
            output_wb.close()
            logger.info("번역된 파일 저장 완료")
            
            # 통계 출력
            total_duration = (datetime.now() - self.start_time).total_seconds()
            logger.info("=" * 80)
            logger.info("번역 완료!")
            logger.info(f"총 소요 시간: {total_duration:.2f}초")
            logger.info(f"총 처리 셀 수: {self.total_cells}")
            logger.info(f"번역된 셀 수: {self.translated_count}")
            logger.info(f"건너뛴 셀 수: {self.skipped_count}")
            logger.info(f"번역 오류 수: {self.error_count}")
            logger.info(f"출력 파일: {output_file}")
            
            print("\n" + "=" * 50)
            print("번역 완료!")
            print(f"총 소요 시간: {total_duration:.2f}초")
            print(f"총 처리 셀 수: {self.total_cells}")
            print(f"번역된 셀 수: {self.translated_count}")
            print(f"건너뛴 셀 수: {self.skipped_count}")
            if self.error_count > 0:
                print(f"번역 오류 수: {self.error_count}")
            print(f"출력 파일: {output_file}")
            
        except InterruptedError as e:
            # 번역 중지 요청 처리
            logger.warning(f"번역 중지됨: {str(e)}")
            if self.should_stop:
                # 이미 번역된 내용이 있다면 저장
                logger.info("번역 중지됨 - 부분적으로 번역된 내용이 있을 수 있습니다.")
                raise  # 상위로 전파하여 GUI에서 처리
            
        except Exception as e:
            logger.critical(f"치명적 오류 발생: {str(e)}", exc_info=True)
            logger.debug(f"전체 스택 트레이스:\n{traceback.format_exc()}")
            raise


def main():
    """메인 함수"""
    logger.info("프로그램 시작")
    logger.debug(f"명령줄 인자: {sys.argv}")
    
    if len(sys.argv) < 2:
        print("사용법: python translate_excel.py <엑셀파일경로> [출력파일경로]")
        print("\n예시:")
        print("  python translate_excel.py input.xlsx")
        print("  python translate_excel.py input.xlsx output.xlsx")
        logger.warning("명령줄 인자 부족")
        sys.exit(1)
    
    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    
    try:
        logger.info("번역기 인스턴스 생성")
        translator = ExcelTranslator(debug_mode=False)  # 성능 향상을 위해 기본값 False
        translator.translate_excel(input_file, output_file)
        logger.info("프로그램 정상 종료")
    except KeyboardInterrupt:
        logger.warning("사용자에 의해 중단됨")
        print("\n\n프로그램이 중단되었습니다.")
        sys.exit(130)
    except Exception as e:
        logger.critical(f"프로그램 오류: {str(e)}", exc_info=True)
        logger.debug(f"전체 스택 트레이스:\n{traceback.format_exc()}")
        print(f"\n오류 발생: {str(e)}")
        print(f"상세 로그는 logs 폴더를 확인하세요.")
        sys.exit(1)


if __name__ == "__main__":
    main()

